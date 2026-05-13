#!/usr/bin/env python3
"""Build financial-model.xlsx from financial-model.yaml.

Usage:
    python3 build_xlsx.py <financial-model.yaml> <output.xlsx>

Renders a multi-sheet Excel workbook:
    - Summary: round, ask, milestones, runway
    - Projections: monthly or quarterly revenue/headcount/burn
    - Unit Economics: ACV, CAC, LTV, payback, NRR, GRR
    - Cohorts: retention table
    - Market Sizing: TAM/SAM/SOM with top-down and bottom-up
    - Use of Funds: milestone-tied spend
    - Assumptions: register with rationale and proof-points
    - Risks (pre-mortem): top 3 risks with mitigation
    - Triggers: stop/pivot conditions
    - Sensitivity: grid (A+ only)
    - Sources: proof-point and source manifest
"""

from __future__ import annotations

import sys
from pathlib import Path

try:
    import yaml
except ImportError:
    print("ERROR: PyYAML is required. Install with: pip install pyyaml", file=sys.stderr)
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
TITLE_FONT = Font(bold=True, size=14)
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_header(ws, row: int, columns: list[str]) -> None:
    for col_idx, label in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER


def write_row(ws, row: int, values: list) -> None:
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def autosize(ws, min_width: int = 12, max_width: int = 60) -> None:
    for col_letter in (get_column_letter(i) for i in range(1, ws.max_column + 1)):
        max_len = 0
        for cell in ws[col_letter]:
            v = cell.value
            if v is None:
                continue
            length = len(str(v))
            max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))


def section_title(ws, row: int, text: str) -> int:
    ws.cell(row=row, column=1, value=text).font = TITLE_FONT
    return row + 2


def build_summary(ws, model: dict) -> None:
    row = 1
    ws.cell(row=row, column=1, value=f"{model.get('company', 'Company')} — Financial Model").font = Font(bold=True, size=16)
    row += 2
    ws.cell(row=row, column=1, value=f"Stage: {model.get('stage', 'unknown')}").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value=f"Horizon: {model.get('horizon_months', 'n/a')} months").font = Font(italic=True)
    row += 2

    round_data = model.get("round", {}) or {}
    row = section_title(ws, row, "Round")
    write_header(ws, row, ["Field", "Value"])
    row += 1
    for key in ("ask_usd", "expected_runway_months", "next_round_milestone", "next_round_target", "justification"):
        write_row(ws, row, [key, round_data.get(key)])
        row += 1
    row += 1

    burn = model.get("burn", {}) or {}
    row = section_title(ws, row, "Burn & Runway")
    write_header(ws, row, ["Field", "Value"])
    row += 1
    for key in (
        "current_cash_usd",
        "current_gross_burn_usd_monthly",
        "current_net_burn_usd_monthly",
        "trailing_3mo_net_burn_avg_usd",
        "runway_months",
        "burn_type_used_for_runway",
    ):
        write_row(ws, row, [key, burn.get(key)])
        row += 1

    autosize(ws)


def build_projections(ws, model: dict) -> None:
    write_header(
        ws,
        1,
        [
            "Period",
            "Revenue (low)",
            "Revenue (base)",
            "Revenue (high)",
            "Headcount",
            "Gross burn",
            "Net burn",
            "Customers active",
            "Notes",
        ],
    )
    row = 2
    for p in model.get("projections", []) or []:
        write_row(
            ws,
            row,
            [
                p.get("period"),
                p.get("revenue_low"),
                p.get("revenue_base"),
                p.get("revenue_high"),
                p.get("headcount"),
                p.get("gross_burn_usd"),
                p.get("net_burn_usd"),
                p.get("customers_active"),
                p.get("notes"),
            ],
        )
        row += 1
    autosize(ws)


def build_unit_economics(ws, model: dict) -> None:
    ue = model.get("unit_economics", {}) or {}
    write_header(ws, 1, ["Metric", "Value", "Confidence / Notes"])
    row = 2
    for key, label, conf_key in [
        ("gross_margin_pct", "Gross margin (%)", "gross_margin_confidence"),
        ("gross_margin_observation_months", "GM observation (months)", None),
        ("cac_blended_usd", "CAC blended (USD)", None),
        ("ltv_usd", "LTV (USD)", None),
        ("ltv_cac_ratio", "LTV : CAC", None),
        ("payback_months", "Payback (months)", None),
        ("nrr_pct", "NRR (%)", None),
        ("grr_pct", "GRR (%)", None),
        ("sales_efficiency_magic_number", "Magic number", None),
    ]:
        conf = ue.get(conf_key) if conf_key else None
        write_row(ws, row, [label, ue.get(key), conf])
        row += 1

    row += 1
    section_title(ws, row, "CAC by channel")
    row += 2
    write_header(ws, row, ["Channel", "CAC (USD)", "Monthly acquisitions", "Proof point"])
    row += 1
    for c in ue.get("cac_by_channel", []) or []:
        write_row(
            ws,
            row,
            [c.get("channel"), c.get("cac_usd"), c.get("monthly_acquisitions"), c.get("proof_point")],
        )
        row += 1
    autosize(ws)


def build_cohorts(ws, model: dict) -> None:
    cohorts = model.get("cohorts", {}) or {}
    ws.cell(row=1, column=1, value=f"Definition: {cohorts.get('definition', '')}").font = Font(italic=True)
    ws.cell(row=2, column=1, value=f"Intervals: {cohorts.get('intervals_measured', [])}")
    ws.cell(row=3, column=1, value=f"Retention type: {cohorts.get('retention_type', '')}")

    write_header(ws, 5, ["Cohort", "Acquired", "Retained M+6", "Retained M+12", "Proof point"])
    row = 6
    for c in cohorts.get("cohort_data", []) or []:
        write_row(
            ws,
            row,
            [
                c.get("cohort"),
                c.get("acquired"),
                c.get("retained_m6"),
                c.get("retained_m12"),
                c.get("proof_point"),
            ],
        )
        row += 1
    autosize(ws)


def build_market(ws, model: dict) -> None:
    market = model.get("market_sizing", {}) or {}
    write_header(ws, 1, ["Scope", "Top-down (USD)", "Bottom-up (USD)", "Convergence", "Confidence"])
    row = 2
    for scope_key in ("tam", "sam", "som"):
        scope = market.get(scope_key, {}) or {}
        write_row(
            ws,
            row,
            [
                scope_key.upper(),
                scope.get("top_down_usd"),
                scope.get("bottom_up_usd"),
                scope.get("convergence_note") or scope.get("restriction_from_tam"),
                scope.get("confidence"),
            ],
        )
        row += 1
    autosize(ws)


def build_use_of_funds(ws, model: dict) -> None:
    write_header(ws, 1, ["Category", "Amount (USD)", "Milestone", "Expected outcome", "Proof point"])
    row = 2
    for u in model.get("use_of_funds", []) or []:
        write_row(
            ws,
            row,
            [
                u.get("category"),
                u.get("amount_usd"),
                u.get("milestone"),
                u.get("expected_outcome"),
                u.get("proof_point"),
            ],
        )
        row += 1
    autosize(ws)


def build_assumptions(ws, model: dict) -> None:
    write_header(ws, 1, ["ID", "Statement", "Rationale", "Proof point", "Risk if wrong"])
    row = 2
    for a in model.get("assumptions", []) or []:
        write_row(
            ws,
            row,
            [
                a.get("id"),
                a.get("statement"),
                a.get("rationale"),
                a.get("proof_point"),
                a.get("risk_if_wrong"),
            ],
        )
        row += 1
    autosize(ws)


def build_risks(ws, model: dict) -> None:
    write_header(ws, 1, ["Reason", "Likelihood", "Impact", "Mitigation"])
    row = 2
    for r in model.get("risks", []) or []:
        write_row(
            ws,
            row,
            [r.get("reason"), r.get("likelihood"), r.get("impact"), r.get("mitigation")],
        )
        row += 1
    autosize(ws)


def build_triggers(ws, model: dict) -> None:
    write_header(ws, 1, ["Bet", "Target", "Pivot condition", "Pivot action"])
    row = 2
    for t in model.get("triggers", []) or []:
        write_row(
            ws,
            row,
            [
                t.get("bet"),
                t.get("target"),
                t.get("pivot_condition"),
                t.get("pivot_action"),
            ],
        )
        row += 1
    autosize(ws)


def build_sensitivity(ws, model: dict) -> None:
    sens = model.get("sensitivity", {}) or {}
    if not sens.get("enabled"):
        ws.cell(row=1, column=1, value="Sensitivity analysis disabled for this stage (FN-14).").font = Font(italic=True)
        return

    write_header(ws, 1, ["Driver", "Downside", "Base", "Upside"])
    row = 2
    for d in sens.get("drivers", []) or []:
        write_row(ws, row, [d.get("name"), d.get("downside"), d.get("base"), d.get("upside")])
        row += 1

    row += 1
    section_title(ws, row, "Outcomes")
    row += 2
    write_header(ws, row, ["Scenario", "Runway (months)", "ARR @ 12mo (USD)"])
    row += 1
    for o in sens.get("outcomes", []) or []:
        write_row(ws, row, [o.get("scenario"), o.get("runway_months"), o.get("arr_12mo_usd")])
        row += 1
    autosize(ws)


def build(model_path: Path, output_path: Path) -> int:
    if not model_path.is_file():
        print(f"ERROR: model not found: {model_path}", file=sys.stderr)
        return 1

    model = yaml.safe_load(model_path.read_text()) or {}

    wb = Workbook()
    # The default sheet becomes Summary
    summary = wb.active
    summary.title = "Summary"
    build_summary(summary, model)

    for name, builder in [
        ("Projections", build_projections),
        ("Unit Economics", build_unit_economics),
        ("Cohorts", build_cohorts),
        ("Market Sizing", build_market),
        ("Use of Funds", build_use_of_funds),
        ("Assumptions", build_assumptions),
        ("Risks (pre-mortem)", build_risks),
        ("Triggers", build_triggers),
        ("Sensitivity", build_sensitivity),
    ]:
        ws = wb.create_sheet(title=name)
        builder(ws, model)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Wrote {output_path}")
    return 0


def main(argv: list[str]) -> int:
    if len(argv) != 3:
        print(__doc__, file=sys.stderr)
        return 1
    return build(Path(argv[1]), Path(argv[2]))


if __name__ == "__main__":
    sys.exit(main(sys.argv))
